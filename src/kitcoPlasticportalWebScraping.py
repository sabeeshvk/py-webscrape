from selenium import webdriver
from selenium.webdriver import Chrome
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import re
import pandas as pd
import os
import requests
import pyodbc
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager
from datetime import date 
import time
driver = webdriver.Chrome(
    "C:/Users/h366319/AppData/Local/Programs/Python/chromedriver.exe")
path = 'C:/workspace/Reports/Kitco and Plasticportal.xlsx'
workbook= load_workbook(path)
next_row_Gold = workbook['Gold'].max_row
next_row_Silver = workbook['Silver'].max_row
next_row_Platinum = workbook['Platinum'].max_row
print ("Starting row: " + str(next_row_Gold))
listURLS = ["https://online.kitco.com/refining/gold-silver.html"]    
length = len(listURLS)
driver.get(listURLS[0])   
soup = BeautifulSoup(requests.get(listURLS[0]).text, "lxml")
table = soup.findAll("table", attrs={"class": "table_selling"})
df = pd.read_html(str(table), header=0)
df[0].drop(df[0].tail(1).index,inplace=True) 
df[1].drop(df[1].tail(1).index,inplace=True) 
df[2].drop(df[2].tail(1).index,inplace=True) 
df[0].insert(4,'Date',date.today())
df[1].insert(4,'Date',date.today())
df[2].insert(4,'Date',date.today())
print(df[0])
print(df[1])
print(df[2])
with pd.ExcelWriter(path, engine='openpyxl', mode='a') as writer:#pylint: disable=abstract-class-instantiated
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)   
    df[0].to_excel(writer,
                    sheet_name="Gold",
                    startrow = next_row_Gold,
                    header=False,
                    index=False)
    df[1].to_excel(writer,
                    sheet_name="Silver",
                    startrow = next_row_Silver,
                    header=False,
                    index=False)
    df[2].to_excel(writer,
                    sheet_name="Platinum",
                    startrow = next_row_Platinum,
                    header=False,
                    index=False)               
#   next_row = next_row + (int((df[0].size)/5))                 
print ("Next Row =" +  str(next_row_Gold))
writer.save
time.sleep(4) 
next_row_Plasticportal = workbook['Plasticportal'].max_row
url = "https://www.plasticportal.eu/en/cenove-reporty/"
driver.get(url)   
soup = BeautifulSoup(requests.get(url).text, "lxml")
table = soup.findAll("table", attrs={"class": "report-table"})
df = pd.read_html(str(table), header=0)
df[0].insert(2,'Date',date.today())
with pd.ExcelWriter(path, engine='openpyxl', mode='a') as writer:#pylint: disable=abstract-class-instantiated
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)   
        df[0].to_excel(writer,
                       sheet_name="Plasticportal",
                       startrow = next_row_Plasticportal,
                       header=False,
                       index=False)             
 #   next_row = next_row + (int((df[0].size)/5))                 
writer.save                      
driver.quit()
