from selenium import webdriver
from selenium.webdriver import Chrome
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import re
import pandas as pd
import os
import requests
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager
from datetime import date 
import time
from selenium.webdriver.chrome.options import Options
from contextlib import suppress
from translate import Translator
path = 'C:/workspace/Reports/soliaoMaster.xlsx'
chrome_path= "C:/Users/h366319/AppData/Local/Programs/Python/chromedriver.exe"
workbook = load_workbook(path)
next_row = workbook['Soliao'].max_row
#url = "https://www.soliao.com/mall-list.html?brandMall=true&qt=&genericName=&supplierName=&flameRating=&fillersName=&fillersContent=&weatherResistance=&hdt=&hardness=&productName=&safeGrade=&hasStock=&mvr=&method=&use=&feature=&origin=false&uv=&hasPrice=&certificate=&searchv=&sortedBySalesVolume=false&sortedByPrice=false&sortedWay=&supplierType=&deliveryAddress=&supplierLevel=&pageNo=1"
url="https://www.soliao.com/mall-list2.html?"
options = Options()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
options.add_argument('disable-infobars')
options.add_argument('--disable-gpu')
options.add_argument('--v=1')
options.add_experimental_option('useAutomationExtension', False)
options.headless = True
prefs = {
  "translate_whitelists": {"zh-CN":"en"},
  "translate":{"enabled":"true"}
}
options.add_experimental_option("prefs", prefs)
products=[] 
prices=[]
manf=[]
driver = webdriver.Chrome(
    "C:/Users/h366319/AppData/Local/Programs/Python/chromedriver.exe")
currencyConvURL="https://www.xe.com/currencyconverter/convert/?Amount=1&From=CNY&To=USD"
driver.get(currencyConvURL)
currencyValClass=driver.find_element_by_class_name('converterresult-toAmount')
currentcurrency = float(currencyValClass.text)
driver.quit()
length=84
for i in range(length):
 with suppress(Exception):
  updatedURL = url+"&currentIndex=" + str(i+1)
  print (updatedURL)
  driver=webdriver.Chrome(executable_path=chrome_path, options=options)
  driver.implicitly_wait(40)
  driver.get(updatedURL)
  content = driver.page_source
  soup = BeautifulSoup(driver.page_source, 'html.parser')
  for a in soup.findAll('div',attrs={'class':'item-text'}):
    name=a.find('div', attrs={'class':'name-text'})
    manfdata=a.find('span', attrs={'class':'generic-supplier-text'})
    translator= Translator(to_lang="en")
    translation = translator.translate(manfdata)
    price=a.find('div', attrs={'class':'price-text'})
    if (name is not None):
      products.append(name.text.strip('\n'))
      priceDetails=re.findall('[0-9]+',price.text.strip('\n'))
      if priceDetails:
        prices.append(float(priceDetails[0]))
      else:
        prices.append(0)
      manf.append(manfdata.text.strip('\n'))    
df = pd.DataFrame({'Product Name':products,'manf Name':manf,'Price':prices})
df["Price"]=currentcurrency*df["Price"]
df["Price"]=df["Price"]/1000
df.insert(3,'Date',date.today())
print(df)   
with pd.ExcelWriter(path, engine='openpyxl', mode='a') as writer:#pylint: disable=abstract-class-instantiated
  writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
  df.to_excel(writer,sheet_name="Soliao",startrow = next_row,header=False,index=False)
  next_row = next_row + (int((df.size)/4))      
  writer.save    
driver.quit()
	