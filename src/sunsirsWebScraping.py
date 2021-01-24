from selenium import webdriver
from selenium.webdriver import Chrome
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import re
import pandas as pd
import os
import pyodbc
import requests
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager
from datetime import date 
import time
from sqlalchemy import create_engine
import urllib
url = "http://www.sunsirs.com/uk/sectors-15.html"
driver = webdriver.Chrome(
    "C:/Users/h366319/AppData/Local/Programs/Python/chromedriver.exe")
path = 'C:/workspace/Reports/sunsirs.xlsx'
workbook = load_workbook(path)
next_row = workbook['sunsirs'].max_row
listURLS = ["http://www.sunsirs.com/uk/sectors-11.html", "http://www.sunsirs.com/uk/sectors-12.html", "http://www.sunsirs.com/uk/sectors-13.html","http://www.sunsirs.com/uk/sectors-14.html","http://www.sunsirs.com/uk/sectors-15.html","http://www.sunsirs.com/uk/sectors-16.html","http://www.sunsirs.com/uk/sectors-17.html"]    
length = len(listURLS)
print(length)
currencyConvURL="https://www.xe.com/currencyconverter/convert/?Amount=1&From=CNY&To=USD"
driver.get(currencyConvURL)
currencyValClass=driver.find_element_by_class_name('converterresult-toAmount')
currentcurrency = float(currencyValClass.text)
for i in range(length):    
    driver.get(listURLS[i])   
    soup = BeautifulSoup(requests.get(listURLS[i]).text, "lxml")
    table = soup.find("table", attrs={"class": "dalistbg"})
    df = pd.read_html(str(table), header=0)
    df[0].drop(df[0].columns[[2, 4]], axis=1, inplace=True)
    df[0].insert(3,'Date',date.today())
    colName=df[0].columns[2]  
    df[0] = df[0].astype({colName: float})
    df[0][colName]=currentcurrency*df[0][colName]
    print(df[0])
    with pd.ExcelWriter(path, engine='openpyxl', mode='a') as writer:#pylint: disable=abstract-class-instantiated
     writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)   
     df[0].to_excel(writer,
                       sheet_name="sunsirs",
                       startrow = next_row,
                       header=False,
                       index=False)    
     next_row = next_row + (int((df[0].size)/4))                 
    print ("Next Row =" +  str(next_row))
    writer.save
    time.sleep(4)                  
driver.quit()
