from selenium import webdriver
from selenium.webdriver import Chrome
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import re
import pandas as pd
import os
import pyodbc
import requests
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager
from datetime import date 
import time
from sqlalchemy import create_engine
import urllib
import pylint
from datetime import date
import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
url = "https://worldfreightrates.com/en/freight"
CHROMEDRIVER_PATH="C:/Users/h366319/AppData/Local/Programs/Python/chromedriver.exe"
#options = Options()
#options.headless = True
driver = webdriver.Chrome(CHROMEDRIVER_PATH)
excelpath = 'C:/workspace/Reports/worldfreightrates.xlsx'
workbook = load_workbook(excelpath)
#next_row = workbook['Source'].max_row
#next_row1 = workbook['freightrates'].max_row
listURLS = ["https://worldfreightrates.com/en/freight"]  
arrPortsFrom=["Mumbai","Shanghai","Penang","Miami","Veracruz","Victoria","Bremerhaven","Bremerhaven","Constantza","Rouen","Tokyo","Venice","Szczecin","Zeebrugge","Varna","Varna","Limerick","Sydney","Westport","Zeeland Seaports","Hong Kong","Taichung","Taichung","Map Ta Phut","Vysotsk","Varna","Piran","Vigo","Sedef","Tilbury","Vitoria","Rio Grande","Vlore","Volos","Singapore"]  
#arrPortsFrom=["Mumbai","Shanghai","Penang"]
#arrPortsTo=["Mumbai","Shanghai","Penang"]
arrPortsTo=["Mumbai","Shanghai","Penang","Miami","Veracruz","Victoria","Bremerhaven","Bremerhaven","Constantza","Rouen","Tokyo","Venice","Szczecin","Zeebrugge","Varna","Varna","Limerick","Sydney","Westport","Zeeland Seaports","Hong Kong","Taichung","Taichung","Map Ta Phut","Vysotsk","Varna","Piran","Vigo","Sedef","Tilbury","Vitoria","Rio Grande","Vlore","Volos","Singapore"]
lengthArrPortsFrom = len(arrPortsFrom)
lengthArrPortsTo = len(arrPortsTo)

dict_rows = {}
desc = []
prange = []
avg = []
change = [] 
dateToday = []
desc1 = []
price1 =[]
Open1 =[]
High1 =[]
Low1 = []
Vol1 =[]
Change1= []
arrFrm=[]
arrTo=[]
arrVal=[]
driver.get(listURLS[0])
for i in range(lengthArrPortsFrom):
    for j in range(lengthArrPortsTo):  
        driver.refresh()
        inputFrom = driver.find_element_by_id("fromNameOcean")
        inputFrom.send_keys(arrPortsFrom[i])  
        time.sleep(2)
        inputFrom.send_keys(Keys.ARROW_DOWN) 
        inputFrom.send_keys(Keys.RETURN) 
        inputTo = driver.find_element_by_id("toNameOcean")
        inputTo.send_keys(arrPortsTo[j])   
        time.sleep(2)
        inputTo.send_keys(Keys.ARROW_DOWN) 
        inputTo.send_keys(Keys.RETURN)    
        inputCommodityVal=driver.find_element_by_id("commodityValueOcean")
        inputCommodityVal.send_keys("$100,000")        
        driver.find_element_by_id("dk_container_commodityNameOcean").click()
        commodity = driver.find_element_by_link_text('Chemicals')    
        commodity.click()   
        time.sleep(2)         
        driver.find_element_by_xpath("//div[@id='OceanSection']/a").click()
        time.sleep(1)     
        val=driver.find_element_by_class_name('result').text    
        arrFrm.append(arrPortsFrom[i])
        arrTo.append(arrPortsTo[j])
        arrVal.append(val)
        print(val)
dict_rows = {'From Port':arrFrm, 'To Port':arrTo, 'Rate': arrVal }
df = pd.DataFrame(dict_rows) 
print(df)
driver.quit


