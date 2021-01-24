from selenium import webdriver
from selenium.webdriver import Chrome
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import re
import pandas as pd
import os
import requests
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager
from datetime import date 
import time
import sys
from selenium.webdriver.chrome.options import Options
from contextlib import suppress
#driver = webdriver.Chrome("C:/Users/h366319/AppData/Local/Programs/Python/Python39-32/Lib/site-packages/chromedriver.exe")
SoliaoSourceLoc = "C:/workspace/Requests/SoliaoSourceLIst.xlsx"
excelOutputPath = 'C:/workspace/Reports/Soliao195.xlsx'
chrome_path= "C:/workspace/scrape/chromedriver.exe"
workbookSource = load_workbook(SoliaoSourceLoc)
sheet_obj = workbookSource.active 
m_row = sheet_obj.max_row   
options = Options()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
options.add_argument('disable-infobars')
options.add_argument('--disable-gpu')
options.add_argument('--v=1')
options.add_experimental_option('useAutomationExtension', False)
options.headless = True
prefs = {
  "translate_whitelists": {"zh-CN":"en"},
  "translate":{"enabled":"true"}
}
options.add_experimental_option("prefs", prefs)
products=[] 
pricesStarting=[]
pricesEnd=[]
length=m_row + 1
#length=3
driver = webdriver.Chrome(
    "C:/Users/h366319/AppData/Local/Programs/Python/chromedriver.exe")
currencyConvURL="https://www.xe.com/currencyconverter/convert/?Amount=1&From=CNY&To=USD"
driver.get(currencyConvURL)
currencyValClass=driver.find_element_by_class_name('converterresult-toAmount')
currentcurrency = float(currencyValClass.text)
driver.quit()
for i in range(2, length):
  with suppress(Exception):
    cell_obj = sheet_obj.cell(row = i, column = 1)
    updatedURL = str(cell_obj.value)
    print (updatedURL)   
    driver=webdriver.Chrome(executable_path=chrome_path, options=options)
    driver.implicitly_wait(40)
    driver.get(updatedURL)
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    name=soup.find('h1', attrs={'class':'product-name-h'})
    price=soup.find('span', attrs={'class':'priceSpan'})   
    if (name is not None):
      products.append(name.text.strip('\n'))
      pricesDetails=re.split(r"[~]",(price.text.strip('\n')))
      pricesStarting.append(float(pricesDetails[0]))
      pricesEnd.append(float(pricesDetails[1]))
  df = pd.DataFrame({'Product Name':products,'PriceStarting':pricesStarting})
  df["PriceStarting"]=currentcurrency*df["PriceStarting"]  
  df["PriceStarting"]=df["PriceStarting"]/1000
  df.insert(2,'Date',date.today())
print(df)  
next_row=2
driver.quit()
workbook = load_workbook(excelOutputPath)
next_row = workbook['Soliao'].max_row
with pd.ExcelWriter(excelOutputPath, engine='openpyxl', mode='a') as writer:#pylint: disable=abstract-class-instantiated
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)   
    df.to_excel(writer,sheet_name="Soliao",startrow = next_row,header=False,index=False)    
    next_row = next_row + (int((df.size)/4))                 
    print ("Next Row =" +  str(next_row))
    writer.save