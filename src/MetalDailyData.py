from selenium import webdriver
from selenium.webdriver import Chrome
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import re
import pandas as pd
import os
import pyodbc
import requests
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager
from datetime import date 
import time
from sqlalchemy import create_engine
import urllib
import pylint
from datetime import date
import time
from selenium.webdriver.chrome.options import Options
url = "https://www.metal.com/price"
CHROMEDRIVER_PATH="C:/Users/h366319/AppData/Local/Programs/Python/chromedriver.exe"
#options = Options()
#options.headless = True
driver = webdriver.Chrome(CHROMEDRIVER_PATH)
excelpath = 'C:/workspace/Reports/Metal.xlsx'
workbook = load_workbook(excelpath)
next_row = workbook['Metal1'].max_row
next_row1 = workbook['Metal2'].max_row
listURLS = ["https://www.metal.com/price","https://www.metal.com/price/Non-Ferrous%20Metals/Copper","https://www.metal.com/price/Non-Ferrous%20Metals/Zinc","https://www.metal.com/price/Non-Ferrous%20Metals/Nickel","https://www.metal.com/price/Non-Ferrous%20Metals/Lead","https://www.metal.com/price/Non-Ferrous%20Metals/Tin"]    
length = len(listURLS)
print(length)
dict_rows = {}
#currencyConvURL="https://www.xe.com/currencyconverter/convert/?Amount=1&From=CNY&To=USD"
#driver.get(currencyConvURL)
#driver.implicitly_wait(15)
#try:
    #currencyValClass=driver.find_element_by_class_name('converterresult-toAmount')
#except Exception:
   #pass
#print(currencyValClass.text)
#df = pd.DataFrame(columns=['Price description', 'Price Range', 'Avg.', 'Change', 'Date'])

desc = []
prange = []
avg = []
change = [] 
dateToday = []
desc1 = []
price1 =[]
Open1 =[]
High1 =[]
Low1 = []
Vol1 =[]
Change1= []
for i in range(length):    
    driver.get(listURLS[i])   
    soup = BeautifulSoup(requests.get(listURLS[i]).text, "lxml")
    divTag = soup.find_all("div", attrs={"class": "row___1xJWs close___30tSe"})
    for tag in divTag:       
        innerDivtags = tag.find_all("div", {"class": "item___ku9Fy"})
        materialName=tag.find('a').contents[0]
        if(len(innerDivtags))<5:            
            desc.append (materialName)        
            i=1        
            for innerdivs in innerDivtags:
                if i==1:
                    prange.append(innerdivs.text)
                     
                if i==2:
                    avg.append (innerdivs.text)
                     
                if i==3:
                    change.append (innerdivs.text)
                     
                if i==4:
                    dateToday.append (innerdivs.text)
                     
                i=i+1
        else:
           desc1.append(materialName)
           i=1
           for innerdivs in innerDivtags:
                if i==1:
                    price1.append(innerdivs.text)
                     
                if i==2:
                    Open1.append(innerdivs.text)
                     
                if i==3:
                    High1.append (innerdivs.text)
                     
                if i==4:
                    Low1.append (innerdivs.text)
                     
                if i==5:
                    Vol1.append (innerdivs.text)
                     
                if i==6:
                    Change1.append(innerdivs.text)
                     
                i=i+1

    dict_rows = {'Price description':desc, 'Price range':prange, 'Avg': avg, 'Change' : change, 'Date': dateToday }
    dict_rows1 = {'Price description':desc1, 'Price':price1, 'open': Open1, 'High' : High1, 'Low': Low1,'Vol':Vol1,'Change':Change1 }
    df = pd.DataFrame(dict_rows) 
    df1=pd.DataFrame(dict_rows1)    
    df1.insert(7,'Date',date.today())
    with pd.ExcelWriter(excelpath, engine='openpyxl', mode='a') as writer:#pylint: disable=abstract-class-instantiated
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)   
        df.to_excel(writer,
                       sheet_name="Metal1",
                       startrow = next_row,
                       header=False,
                       index=False)    
        next_row = next_row + 1               
    writer.save
    writer.close
    time.sleep(4)
    with pd.ExcelWriter(excelpath, engine='openpyxl', mode='a') as writer:#pylint: disable=abstract-class-instantiated
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)   
        df1.to_excel(writer,
                       sheet_name="Metal2",
                       startrow = next_row1,
                       header=False,
                       index=False)    
        next_row1 = next_row1 + 1               
    writer.save
    writer.close
    time.sleep(4)
    print(df)
    print(df1)
driver.quit()
